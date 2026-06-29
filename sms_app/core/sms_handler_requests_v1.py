#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS 处理器 - Requests 版本（不使用 Selenium）
高性能、直接 HTTP API 调用实现
"""

import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from bs4 import BeautifulSoup
import json
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

# 导入全局常量配置
from .constants import LOGIN_URL, ACTIVITY_PAGE

# 禁用 SSL 警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


class SMSHandlerRequests:
    """使用 requests 库与 SMS 系统交互（无 Selenium）"""
    
    # 从全局常量中使用URL配置（2026-06-20 改为内网地址）
    LOGIN_URL = LOGIN_URL
    ACTIVITY_PAGE = ACTIVITY_PAGE
    POST_URL = ACTIVITY_PAGE  # POST 使用同一个地址
    
    def __init__(self, excel_path: str = None):
        """初始化处理器
        
        Args:
            excel_path: Excel 文件路径（含学生成绩数据）
        """
        self.session = None
        self.excel_path = excel_path or str(Path.home() / "Upload.xlsx")
        self.class_mapping = {}  # class_id → class_name
        self.club_mapping = {}   # club_id → club_name
        self.mappings_file = Path.home() / ".sms_app" / "mappings.json"
    
    def _init_session(self):
        """初始化请求会话"""
        if not self.session:
            self.session = requests.Session()
            self.session.verify = False
            # 设置超时和重试
            self.session.timeout = 30
        return self.session
    
    def _load_mappings(self):
        """加载班级和学会映射表"""
        try:
            if self.mappings_file.exists():
                with open(self.mappings_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.class_mapping = data.get('classes', {})
                    self.club_mapping = data.get('clubs', {})
                    print(f" 已加载映射表: {len(self.class_mapping)} 个班级, {len(self.club_mapping)} 个学会")
                    return True
        except Exception as e:
            print(f"  加载映射表失败: {e}")
        return False
    
    def login(self, username: str, password: str) -> bool:
        """使用 requests 登入 SMS 系统
        
        Args:
            username: 用户名
            password: 密码
            
        Returns:
            登入是否成功
        """
        try:
            session = self._init_session()
            print(f" 使用 requests 登入...")
            print(f"   用户: {username}")
            
            # 第一步：GET 登入页面（获取初始 cookie）
            print(f"   1⃣  获取登入页面...")
            resp = session.get(self.LOGIN_URL, timeout=15)
            if resp.status_code != 200:
                print(f" 获取登入页面失败: {resp.status_code}")
                return False
            print(f"    页面已获取")
            
            # 第二步：POST 登入凭证
            print(f"   2⃣  提交登入凭证...")
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            resp = session.post(self.LOGIN_URL, data=login_data, timeout=15, allow_redirects=True)
            
            # 检查是否仍在登入页面
            if 'login' in resp.url.lower():
                print(f" 登入失败: 仍停留在登入页面")
                return False
            
            print(f" 登入成功！")
            print(f"   当前 URL: {resp.url}")
            return True
            
        except Exception as e:
            print(f" 登入异常: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_class_id_from_name(self, class_name: str) -> Optional[str]:
        """从班级名称查找班级 ID
        
        Args:
            class_name: 班级名称或简称 (如 "S3A", "J1A")
            
        Returns:
            班级 ID，如果找不到则返回 None
        """
        # 直接尝试从映射表查找（班级 ID → 班级名称）
        for class_id, class_full_name in self.class_mapping.items():
            # 如果是字符串，尝试从括号提取简称
            if isinstance(class_full_name, str):
                # 格式: "初一忠 (J1A)"
                if f"({class_name.upper()})" in class_full_name or class_full_name.endswith(class_name.upper()):
                    return class_id
            else:
                # 如果是对象，查找 'short' 字段
                class_short = class_full_name.get('short', '')
                if class_short.upper() == class_name.upper():
                    return class_id
        
        return None
    
    def _get_students_by_class(self, class_name: str, class_id_sms: str = None) -> List[Dict]:
        """获取指定班级的学生列表
        
        Args:
            class_name: 班级名称（如 "S3A"）
            class_id_sms: SMS 中的班级 ID（如果已知）
            
        Returns:
            学生列表，每项包含 internal_id, student_no, name, class_id 等
        """
        try:
            # 如果没提供班级 ID，尝试从名称查找
            if not class_id_sms:
                class_id_sms = self._get_class_id_from_name(class_name)
                if not class_id_sms:
                    print(f"     无法找到班级 {class_name} 的 ID")
                    return []
            
            # GET 活动页面，选择班级时触发 AJAX 获取学生列表
            params = {'class_id': class_id_sms}
            resp = self.session.get(self.ACTIVITY_PAGE, params=params, timeout=15)
            
            if resp.status_code != 200:
                print(f"     获取班级 {class_name} (ID: {class_id_sms}) 学生列表失败: {resp.status_code}")
                return []
            
            # 解析 HTML
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 找所有学生行 - a 标签带 data-student_id
            students = []
            for link in soup.select('a[data-student_id]'):
                student = {
                    'internal_id': link.get('data-student_id'),
                    'student_no': link.get('data-student_no'),
                    'name': link.get('data-student_name'),
                    'name_cn': link.get('data-student_cname'),
                    'class_name': link.get('data-class_name'),
                    'class_id': link.get('data-class_id'),
                }
                if student['internal_id']:  # 确保有 internal_id
                    students.append(student)
            
            print(f"    班级 {class_name} (ID: {class_id_sms}): 找到 {len(students)} 个学生")
            return students
            
        except Exception as e:
            print(f"    获取班级 {class_name} 学生列表异常: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def _read_excel_data(self) -> Tuple[Dict, Dict]:
        """从 Excel 读取学生成绩数据
        
        Returns:
            (students_by_class, field_mapping)
            students_by_class: {class_short: [(row_idx, student_id, name, award), ...]}
            field_mapping: {column_letter: field_name}
        """
        try:
            print(f" 读取 Excel 数据: {self.excel_path}")
            wb = load_workbook(self.excel_path)
            ws = wb.active
            
            # 从第 1 行读取活动信息
            event_date = ws['A1'].value
            item_code = ws['A2'].value
            print(f"   事件: {event_date}")
            print(f"   项目代码: {item_code}")
            
            # 第 4 行是列标题
            headers = {}
            for col_idx in range(1, 10):  # A-I 列
                cell_value = ws.cell(row=4, column=col_idx).value
                if cell_value:
                    headers[col_idx] = cell_value.lower()
            
            print(f"   列标题: {headers}")
            
            # 读取学生数据（从第 5 行开始）
            students_by_class = {}
            for row_idx in range(5, ws.max_row + 1):
                class_name = ws.cell(row=row_idx, column=2).value
                student_id = ws.cell(row=row_idx, column=3).value
                name = ws.cell(row=row_idx, column=1).value
                category = ws.cell(row=row_idx, column=4).value
                award = ws.cell(row=row_idx, column=5).value
                
                if not student_id or not class_name:
                    continue
                
                if class_name not in students_by_class:
                    students_by_class[class_name] = []
                
                students_by_class[class_name].append({
                    'row_idx': row_idx,
                    'student_id': str(student_id),
                    'name': name,
                    'award': award or '',
                    'category': category or ''
                })
            
            print(f"    读取完成: {sum(len(v) for v in students_by_class.values())} 个学生")
            for cls, students in students_by_class.items():
                print(f"      {cls}: {len(students)} 人")
            
            return students_by_class, headers
            
        except Exception as e:
            print(f" 读取 Excel 异常: {e}")
            import traceback
            traceback.print_exc()
            return {}, {}
    
    def upload_student_scores(self, username: str, password: str, 
                             year: str = "2026", semester: str = "1",
                             date: str = "", item_id: str = "") -> Dict:
        """上传学生成绩（requests 版本）
        
        Args:
            username: SMS 用户名
            password: SMS 密码
            year: 年份
            semester: 学期
            date: 活动日期
            item_id: 项目 ID
            
        Returns:
            {
                'success': bool,
                'uploaded': int,
                'failed': int,
                'total': int,
                'message': str,
                'errors': list
            }
        """
        result = {
            'success': False,
            'uploaded': 0,
            'failed': 0,
            'total': 0,
            'message': '',
            'errors': []
        }
        
        try:
            print(f"\n{'='*60}")
            print(f" SMS 成绩上传 (Requests 版本)")
            print(f"{'='*60}")
            
            # 1⃣ 加载映射表
            self._load_mappings()
            
            # 2⃣ 登入
            print(f"\n 步骤 1/4: 登入")
            if not self.login(username, password):
                result['message'] = " 登入失败"
                print(result['message'])
                return result
            
            # 3⃣ 读取 Excel 数据
            print(f"\n 步骤 2/4: 读取 Excel 数据")
            students_by_class, field_mapping = self._read_excel_data()
            if not students_by_class:
                result['message'] = " Excel 数据为空"
                print(result['message'])
                return result
            
            result['total'] = sum(len(v) for v in students_by_class.values())
            
            # 4⃣ 获取学生列表并构建 POST 数据
            print(f"\n 步骤 3/4: 获取学生列表")
            
            # 从页面获取初始数据（项目列表等）
            resp = self.session.get(self.ACTIVITY_PAGE, timeout=15)
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 提取项目列表（如果没指定 item_id）
            if not item_id:
                first_option = soup.select_one('select#StudentPerformanceM_item_id option[value]:not([value=""])')
                if first_option:
                    item_id = first_option.get('value')
                    print(f"    使用第一个项目: {first_option.text} (ID: {item_id})")
            
            # 构建所有学生的 POST 数据
            post_data = {
                'StudentPerformanceM[year]': year,
                'StudentPerformanceM[semester]': semester,
                'StudentPerformanceM[date]': date,
                'StudentPerformanceM[item_id]': item_id,
            }
            
            total_students = 0
            for class_name, students in students_by_class.items():
                # 获取该班级的学生列表
                sms_students = self._get_students_by_class(class_name)
                
                if not sms_students:
                    print(f"     班级 {class_name} 无法获取学生列表，跳过")
                    for student in students:
                        result['errors'].append(f"无法获取 {class_name} 的学生列表")
                        result['failed'] += 1
                    continue
                
                # 构建学号→SMS学生数据的映射
                sms_student_map = {s['student_no']: s for s in sms_students}
                
                for student in students:
                    student_id = student['student_id']
                    award = student['award']
                    
                    # 查找对应的 SMS 学生
                    if student_id not in sms_student_map:
                        print(f"     未找到学生: {class_name} {student_id}")
                        result['errors'].append(f"未找到 {class_name} {student_id}")
                        result['failed'] += 1
                        continue
                    
                    sms_student = sms_student_map[student_id]
                    internal_id = sms_student['internal_id']
                    
                    # Determine type_of_bonus from category
                    category = student.get('category', '')
                    def _map_category(cat):
                        try:
                            if not cat:
                                return '1'
                            s = str(cat)
                            if '特殊' in s:
                                return '2'
                            if '校外' in s:
                                return '1'
                            return '1'
                        except Exception:
                            return '1'

                    type_val = _map_category(category)

                    # 添加该学生的表单字段
                    post_data[f'StudentPerformanceM[inputperformance][{internal_id}][class_id]'] = sms_student['class_id']
                    post_data[f'StudentPerformanceM[inputperformance][{internal_id}][type_of_bonus]'] = type_val
                    post_data[f'StudentPerformanceM[inputperformance][{internal_id}][mark]'] = '0.00'
                    post_data[f'StudentPerformanceM[inputperformance][{internal_id}][remark]'] = str(award)
                    
                    total_students += 1
                    print(f"    {class_name} {student_id}: {sms_student['name']} → {award}")
            
            if total_students == 0:
                result['message'] = " 没有有效的学生数据"
                print(result['message'])
                return result
            
            # 5⃣ 提交 POST 请求
            print(f"\n 步骤 4/4: 提交数据")
            print(f"    POST 到: {self.POST_URL}")
            print(f"    数据字段数: {len(post_data)}")
            
            resp = self.session.post(self.POST_URL, data=post_data, timeout=30)
            
            if resp.status_code == 200:
                # 检查响应中是否有成功标识
                if '成功' in resp.text or '已创建' in resp.text or 'success' in resp.text.lower():
                    result['success'] = True
                    result['uploaded'] = total_students
                    result['message'] = f" 成功上传 {total_students} 个学生"
                else:
                    # 检查是否有错误提示
                    soup = BeautifulSoup(resp.text, 'html.parser')
                    error_msg = soup.find('div', class_='alert-error')
                    if error_msg:
                        result['message'] = f"  上传可能失败: {error_msg.text}"
                    else:
                        result['success'] = True
                        result['uploaded'] = total_students
                        result['message'] = f" 已提交 {total_students} 个学生"
            else:
                result['message'] = f" POST 请求失败: {resp.status_code}"
                result['errors'].append(f"HTTP {resp.status_code}")
            
            print(result['message'])
            return result
            
        except Exception as e:
            result['message'] = f" 异常: {str(e)}"
            result['errors'].append(str(e))
            print(result['message'])
            import traceback
            traceback.print_exc()
            return result
        
        finally:
            # 清理
            if self.session:
                self.session.close()
                self.session = None
