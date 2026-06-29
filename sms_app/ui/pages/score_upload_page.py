#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
成绩上传页面 - 批量上传学生成绩
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFileDialog, QGroupBox, QTableWidget, QTableWidgetItem, QProgressBar
)
from PyQt6.QtGui import QFont
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from pathlib import Path
from openpyxl import load_workbook
import time

from core.config_manager import ConfigManager


class UploadThread(QThread):
    """后台线程执行上传（支持自动重新登入）"""
    progress_updated = pyqtSignal(int)
    upload_finished = pyqtSignal(bool, str)
    log_message = pyqtSignal(str, str)
    
    def __init__(self, excel_path: str, username: str, password: str, session=None):
        super().__init__()
        self.excel_path = excel_path
        self.username = username
        self.password = password
        self.session = session  # 从启动时保存的 session
    
    def run(self):
        """
        执行上传流程：
        1. 读取 Excel 数据
        2. 调用 SMS 处理器上传（会自动处理登入失败/cookie过期）
        """
        try:
            def emit_log(level: str, message: str):
                self.log_message.emit(level, message)

            self.progress_updated.emit(10)
            
            # 第1步：读取 Excel 文件
            emit_log('info', f"[上传线程] 读取 Excel 文件: {self.excel_path}")
            from openpyxl import load_workbook
            
            wb = load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            
            # 先读取日期（从 B1）
            date_val = ws.cell(row=1, column=2).value  # B1
            if isinstance(date_val, str):
                date_str = date_val
            else:
                try:
                    date_str = date_val.strftime('%Y-%m-%d')
                except:
                    date_str = '2026-01-01'

            # 读取活动代码（从 B2）
            activity_code_val = ws.cell(row=2, column=2).value  # B2
            activity_code = str(activity_code_val).strip() if activity_code_val else ''
            if not activity_code:
                activity_code = 'ACA CMO207'
                emit_log('warning', "[上传线程] B2 未读取到活动代码，降级使用 ACA CMO207")
            else:
                emit_log('info', f"[上传线程] 读取到活动代码: {activity_code}")
            
            # 读取数据（从第5行开始）
            scores_data = []
            for row_num in range(5, ws.max_row + 1):
                raw_student_id = ws.cell(row=row_num, column=3).value
                if raw_student_id is None or str(raw_student_id).strip() == "":
                    student_id = ""
                elif isinstance(raw_student_id, (int, float)):
                    student_id = str(int(raw_student_id))
                else:
                    student_id = str(raw_student_id).strip()

                cell_values = {
                    'name': ws.cell(row=row_num, column=1).value,      # Column 1: 姓名
                    'class': ws.cell(row=row_num, column=2).value,     # Column 2: 班级
                    'student_id': student_id,                          # Column 3: 学号
                    'category': ws.cell(row=row_num, column=4).value,  # Column 4: 类别
                    'remarks': ws.cell(row=row_num, column=5).value,   # Column 5: 奖项/备注
                    'english_name': ws.cell(row=row_num, column=6).value, # Column 6: 英文名
                }
                
                # 如果至少有学号和班级，则视为有效行
                if cell_values['student_id'] and cell_values['class']:
                    scores_data.append(cell_values)
            
            wb.close()
            
            emit_log('info', f"[上传线程] 读取到 {len(scores_data)} 条有效数据")
            emit_log('info', f"[上传线程] 日期: {date_str}, 活动代码: {activity_code}")
            
            if len(scores_data) == 0:
                self.upload_finished.emit(False, "❌ 未找到有效的学生数据")
                return
            
            self.progress_updated.emit(20)
            
            # 第2步：创建 SMS 处理器并上传
            emit_log('info', "[上传线程] 创建 SMS 处理器...")
            from core.sms_handler import SMSHandler
            
            handler = SMSHandler(headless=False)
            
            self.progress_updated.emit(30)
            
            emit_log('info', "[上传线程] 开始上传数据（带自动重新登入机制）...")
            
            # 调用上传方法，传入 session（如果存在）
            result = handler.upload_student_scores(
                username=self.username,
                password=self.password,
                scores_data=scores_data,
                date=date_str,
                activity_code=activity_code,
                session=self.session,  # 传入启动时保存的 session
                max_retries=3,  # 最多重试3次
                retry_delay=2,  # 每次重试延迟2秒
                log_callback=emit_log,
            )
            
            # 根据结果更新进度
            if result['success']:
                self.progress_updated.emit(90)
                time.sleep(0.5)
                self.progress_updated.emit(100)
                
                message = f"✅ 上传成功！({result['uploaded']}/{result['total']} 条)"
                emit_log('success', f"[上传线程] {message}")
                self.upload_finished.emit(True, message)
            else:
                # 部分成功或完全失败
                self.progress_updated.emit(100)
                
                if result['uploaded'] > 0:
                    message = (f"⚠️  部分上传成功\n"
                             f"成功: {result['uploaded']}, 失败: {result['failed']}\n"
                             f"详情: {result['message']}")
                    if result.get('errors'):
                        message += f"\n未找到: {', '.join(result['errors'])}"
                else:
                    message = f"❌ {result['message']}"
                    if result.get('errors'):
                        message += f"\n未找到: {', '.join(result['errors'])}"
                
                emit_log('error' if result['uploaded'] == 0 else 'warning', f"[上传线程] {message}")
                self.upload_finished.emit(result['uploaded'] > 0, message)
            
        except Exception as e:
            self.log_message.emit('error', f"[上传线程] 异常: {e}")
            import traceback
            traceback.print_exc()
            self.upload_finished.emit(False, f"❌ 上传异常: {str(e)}")


class ScoreUploadPage:
    def __init__(self, console, get_session_callback=None):
        self.console = console
        self.widget = None
        self.selected_file = None
        self.config = ConfigManager()
        self.upload_thread = None
        self.get_session_callback = get_session_callback  # 获取 session 的回调函数
        self._create_ui()
    
    def _create_ui(self):
        """创建 UI"""
        self.widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # 标题
        title = QLabel("成绩上传")
        title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # 文件选择区
        file_group = QGroupBox("选择文件")
        file_layout = QVBoxLayout()
        file_layout.setSpacing(10)
        
        file_btn_layout = QHBoxLayout()
        
        self.file_label = QLabel("未选择文件")
        self.file_label.setStyleSheet("background-color: #3c3c3c; padding: 10px; border: 1px solid #3e3e42;")
        
        select_btn = QPushButton("📂 选择 Excel 文件")
        select_btn.setMaximumWidth(150)
        select_btn.clicked.connect(self.select_file)
        
        download_btn = QPushButton("⬇️  下载模板")
        download_btn.setMaximumWidth(150)
        download_btn.clicked.connect(self.download_template)
        
        file_btn_layout.addWidget(self.file_label, 1)
        file_btn_layout.addWidget(select_btn)
        file_btn_layout.addWidget(download_btn)
        
        file_layout.addLayout(file_btn_layout)
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)
        
        # 预览表格
        preview_label = QLabel("📋 成绩数据预览")
        preview_label.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        layout.addWidget(preview_label)
        
        self.preview_table = QTableWidget()
        self.preview_table.setColumnCount(6)
        self.preview_table.setHorizontalHeaderLabels(["班级", "学号", "姓名", "类别", "备注", "英文名"])
        self.preview_table.setMaximumHeight(150)
        layout.addWidget(self.preview_table)
        
        # 进度条
        progress_label = QLabel("上传进度")
        progress_label.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        layout.addWidget(progress_label)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setMaximumHeight(20)
        layout.addWidget(self.progress_bar)
        
        # 上传按钮
        btn_layout = QHBoxLayout()
        upload_btn = QPushButton("📤 开始上传")
        upload_btn.setMinimumWidth(120)
        upload_btn.clicked.connect(self.start_upload)
        
        btn_layout.addWidget(upload_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        layout.addStretch()
        
        self.widget.setLayout(layout)
    
    def get_input_widget(self):
        """返回输入区 widget"""
        return self.widget
    
    def select_file(self):
        """选择文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            None,
            "选择 Excel 文件",
            "",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if file_path:
            self.selected_file = file_path
            self.file_label.setText(f"✓ {file_path}")
            self.console.log_success(f"已选择文件: {file_path}")
            self._load_preview(file_path)
    
    def _load_preview(self, file_path: str):
        """加载 Excel 预览"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # 清空表格
            self.preview_table.setRowCount(0)
            
            # 读取第4行作为标题（如果存在）
            headers = []
            for col in range(1, 7):
                cell_value = ws.cell(row=4, column=col).value
                if cell_value:
                    headers.append(str(cell_value))

            if headers:
                self.preview_table.setColumnCount(len(headers))
                self.preview_table.setHorizontalHeaderLabels(headers)

            # 读取数据行（从第5行开始）
            row_idx = 0
            for excel_row in range(5, min(ws.max_row + 1, 25)):  # 最多显示20行
                col_idx = 0
                row_data = []
                for col in range(1, 7):
                    cell = ws.cell(row=excel_row, column=col)
                    row_data.append(str(cell.value) if cell.value else "")

                if any(row_data):  # 如果该行不为空
                    self.preview_table.insertRow(row_idx)
                    for i, data in enumerate(row_data):
                        self.preview_table.setItem(row_idx, i, QTableWidgetItem(data))
                    row_idx += 1
            
            wb.close()
            self.console.log_info(f"已加载 {row_idx} 行数据", "#4ec9b0")
        except Exception as e:
            self.console.log_error(f"加载文件失败: {str(e)}")
    
    def download_template(self):
        """下载模板"""
        try:
            import sys
            # 改用動態產生範本，避免二進位檔需要管理
            save_path, _ = QFileDialog.getSaveFileName(
                None,
                "保存模板",
                "template.xlsx",
                "Excel Files (*.xlsx)"
            )

            if save_path:
                try:
                    from openpyxl import Workbook

                    wb = Workbook()
                    ws = wb.active

                    # 第1、2 行保留給日期與活動代碼（對應現有上傳程式）
                    ws.cell(row=1, column=1, value='date')
                    ws.cell(row=1, column=2, value='2026-01-01')
                    ws.cell(row=2, column=1, value='activity_code')
                    ws.cell(row=2, column=2, value='ACA CMO207')

                    # 第4行為表頭：name, class, studentId, category, award, english_name
                    headers = ['name', 'class', 'studentId', 'category', 'award', 'english_name']
                    for i, h in enumerate(headers, start=1):
                        ws.cell(row=4, column=i, value=h)

                    # 寫入範例一列（空白為可編輯）
                    ws.cell(row=5, column=1, value='張三')
                    ws.cell(row=5, column=2, value='J1A')
                    ws.cell(row=5, column=3, value='26001')
                    ws.cell(row=5, column=4, value='校外学艺')
                    ws.cell(row=5, column=5, value='優異')
                    ws.cell(row=5, column=6, value='ZHANG SAN')

                    wb.save(save_path)
                    wb.close()

                    self.console.log_success(f"[下载模板] 模板已生成并保存到: {save_path}")
                except Exception as e:
                    self.console.log_error(f"[下载模板] 生成模板失败: {str(e)}")
        except Exception as e:
            self.console.log_error(f"[下载模板] 异常: {str(e)}")
    
    def start_upload(self):
        """开始上传"""
        if not self.selected_file:
            self.console.log_warning("请先选择 Excel 文件")
            return
        
        # 获取保存的凭证
        username, password = self.config.get_credentials()
        if not username or not password:
            self.console.log_warning("未找到保存的凭证，请先在【设定】页面保存凭证")
            return
        
        self.console.log_info(f"开始上传: {self.selected_file}", "#dcdcaa")
        self.progress_bar.setValue(0)
        
        # 获取现有的 session（如果有）
        session = None
        if self.get_session_callback:
            try:
                session = self.get_session_callback()
            except:
                pass
        
        # 启动后台线程
        self.upload_thread = UploadThread(self.selected_file, username, password, session=session)
        self.upload_thread.progress_updated.connect(self._update_progress)
        self.upload_thread.log_message.connect(self._on_upload_log)
        self.upload_thread.upload_finished.connect(self._on_upload_finished)
        self.upload_thread.start()
    
    def _update_progress(self, value: int):
        """更新进度条"""
        self.progress_bar.setValue(value)

    def _on_upload_log(self, level: str, message: str):
        """将后台线程日志写入 UI 控制台与日志文件"""
        if level == 'success':
            self.console.log_success(message)
        elif level == 'warning':
            self.console.log_warning(message)
        elif level == 'error':
            self.console.log_error(message)
        else:
            self.console.log_info(message)
    
    def _on_upload_finished(self, success: bool, message: str):
        """上传完成"""
        if success:
            self.console.log_success(message)
        else:
            self.console.log_error(message)
