#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS 系统 - 批量获取所有班级的学生名单
改进版本，支持网络自动选择和指定输出目录
"""

import requests
from bs4 import BeautifulSoup
import json
import time
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

# 禁用 SSL 警告
requests.packages.urllib3.disable_warnings()


class SMSStudentListFetcher:
    """从 SMS 系统获取所有班级的学生名单"""
    
    # 支持的网络接口
    NETWORKS = {
        'intranet': 'http://192.168.0.6/sms/index.php',      # 局域网
        'internet': 'http://sms.chhsban.edu.my/sms/index.php'  # 广域网
    }
    
    def __init__(self, base_url: str = None, data_dir: str = None):
        """初始化获取器
        
        Args:
            base_url: SMS 系统基础 URL，如果为 None 则自动选择
            data_dir: 数据保存目录，默认为 C:\chhsban-acadoc\download_student
        """
        self.base_url = base_url
        self.session = None
        self.class_mapping = {}  # class_id → class_name
        self.all_students = {}   # class_id → [student_list]
        
        # 数据保存目录
        if data_dir is None:
            self.data_dir = Path("C:/chhsban-acadoc/download_student")
        else:
            self.data_dir = Path(data_dir)
        
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        # 日志文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = self.data_dir / f"download_{timestamp}.log"
    
    def log(self, message: str, level: str = "INFO"):
        """记录日志"""
        log_message = f"[{level}] {message}"
        print(log_message)
        
        with open(self.log_file, 'a', encoding='utf-8') as f:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{timestamp}] {log_message}\n")
    
    def _select_network(self) -> bool:
        """自动选择可用的网络接口"""
        if self.base_url:
            return True
        
        self.log("尝试自动选择网络接口...")
        
        # 优先尝试局域网
        for network_name, url in self.NETWORKS.items():
            self.log(f"  尝试 {network_name}...")
            try:
                resp = requests.get(url, timeout=5, verify=False)
                if resp.status_code == 200:
                    self.base_url = url
                    self.log(f"  ✓ 成功连接到 {network_name}: {url}")
                    return True
            except Exception as e:
                self.log(f"  连接失败: {type(e).__name__}")
                continue
        
        self.log("ERROR: 无法连接到任何网络接口", "ERROR")
        return False
    
    def _init_session(self):
        """初始化请求会话"""
        if not self.session:
            self.session = requests.Session()
            self.session.verify = False
            self.session.timeout = 30
        return self.session
    
    def login(self, username: str, password: str) -> bool:
        """登入 SMS 系统
        
        Args:
            username: 用户名
            password: 密码
            
        Returns:
            是否登入成功
        """
        try:
            # 选择网络
            if not self._select_network():
                return False
            
            session = self._init_session()
            self.log(f"开始登入...")
            self.log(f"  用户: {username}")
            
            # 配置 URL
            login_url = f"{self.base_url}?r=site/login"
            activity_page = f"{self.base_url}?r=transaction/studentPerformance/create"
            
            # 第1步：GET 登入页面
            self.log(f"1️⃣  获取登入页面...")
            resp = session.get(login_url, timeout=15)
            if resp.status_code != 200:
                self.log(f"ERROR: 获取登入页面失败: HTTP {resp.status_code}", "ERROR")
                return False
            self.log(f"✓ 页面已获取")
            
            # 第2步：POST 登入凭证
            self.log(f"2️⃣  提交登入凭证...")
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            resp = session.post(login_url, data=login_data, timeout=15, allow_redirects=True)
            
            # 检查是否仍在登入页面
            if 'login' in resp.url.lower():
                self.log(f"ERROR: 登入失败: 错误的用户名或密码", "ERROR")
                return False
            
            self.log(f"✓ 登入成功")
            self.log(f"  当前 URL: {resp.url}")
            
            # 保存活动页面 URL 供后续使用
            self.activity_page = activity_page
            
            # 第3步：获取班级列表（从 select 的 option 中）
            self.log(f"3️⃣  从 option 提取班级列表...")
            if self._fetch_class_list():
                self.log(f"✓ 班级列表已获取 ({len(self.class_mapping)} 个班级)")
            else:
                self.log(f"WARNING: 班级列表获取失败，但继续处理", "WARN")
            
            return True
            
        except Exception as e:
            self.log(f"ERROR: 登入异常: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return False
    
    def _fetch_class_list(self) -> bool:
        """从页面的 SELECT 的 OPTION 提取班级列表
        
        关键：找到正确的班级 SELECT，不是项目 SELECT
        """
        try:
            resp = self.session.get(self.activity_page, timeout=15)
            if resp.status_code != 200:
                self.log(f"ERROR: 获取活动页面失败: HTTP {resp.status_code}", "ERROR")
                return False
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 第1步：查找所有 select 元素，区分项目 SELECT 和班级 SELECT
            all_selects = soup.find_all('select')
            self.log(f"  页面中共有 {len(all_selects)} 个 select 元素")
            
            class_select = None
            item_select = None
            
            for idx, select in enumerate(all_selects):
                # 检查 select 的 name 属性和 id 属性
                name = select.get('name', '')
                select_id = select.get('id', '')
                options = select.find_all('option')
                options_with_value = [opt for opt in options if opt.get('value', '').strip()]
                
                label = f"select#{idx} (name={name}, id={select_id}): {len(options)} 个 option，其中 {len(options_with_value)} 个有 value"
                self.log(f"    {label}")
                
                # 识别项目 SELECT（通常有 2000+ 个 option）
                if 'item_id' in name or len(options_with_value) > 2000:
                    item_select = select
                    self.log(f"      ➜ 这是项目 SELECT", "WARN")
                
                # 识别班级 SELECT（通常有 50-100 个 option）
                if 'class_id' in name or (50 <= len(options_with_value) <= 500):
                    if len(options_with_value) > 30:  # 确保不是太小的 select
                        class_select = select
                        self.log(f"      ➜ 这是班级 SELECT", "WARN")
            
            if not class_select:
                self.log(f"ERROR: 未找到班级选择器", "ERROR")
                return False
            
            self.log(f"✓ 找到班级 SELECT")
            
            # 第2步：从班级 SELECT 中提取班级列表
            self.class_mapping = {}
            options = class_select.find_all('option')
            
            for option in options:
                class_id = option.get('value', '').strip()
                class_name = option.get_text().strip()
                
                # 过滤条件：
                # 1. class_id 不为空
                # 2. class_id 应该是数字（班级 ID 是数字）
                # 3. class_name 不为空
                if class_id and class_name:
                    try:
                        int(class_id)
                        self.class_mapping[class_id] = class_name
                    except ValueError:
                        # 如果不是数字，跳过
                        pass
            
            self.log(f"✓ 从班级 SELECT 提取了 {len(self.class_mapping)} 个班级")
            
            if len(self.class_mapping) == 0:
                self.log(f"WARNING: 班级列表为空，可能需要手动指定班级 ID", "WARN")
                return False
            
            # 显示前几个班级
            for class_id, class_name in list(self.class_mapping.items())[:5]:
                self.log(f"    {class_id}: {class_name}")
            if len(self.class_mapping) > 5:
                self.log(f"    ... 等共 {len(self.class_mapping)} 个班级")
            
            return len(self.class_mapping) > 0
            
        except Exception as e:
            self.log(f"ERROR: 提取班级列表异常: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_students_by_class(self, class_id: str, class_name: str) -> List[Dict]:
        """获取指定班级的所有学生
        
        使用 class_id 和 ajax=student-grid 参数获取学生数据
        需要注意：可能需要 item_id 或其他参数
        
        Args:
            class_id: 班级 ID（从 option value 获取）
            class_name: 班级名称
            
        Returns:
            学生列表
        """
        try:
            # 方式1：尝试只用 class_id 参数
            query_url = f"{self.base_url}?r=transaction/studentPerformance/create"
            params = {
                'class_id': class_id,
                'ajax': 'student-grid'
            }
            
            resp = self.session.get(query_url, params=params, timeout=15)
            
            if resp.status_code != 200:
                self.log(f"  WARNING: 获取班级 {class_name} (ID: {class_id}) 失败: HTTP {resp.status_code}", "WARN")
                return []
            
            # 解析 HTML
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 查找所有学生行 - a 标签带 data-student_id 属性
            students = []
            for link in soup.select('a[data-student_id]'):
                # 关键：只取 data-class_id 与当前 class_id 相匹配的学生
                link_class_id = link.get('data-class_id', '')
                
                if link_class_id == class_id:
                    student = {
                        'internal_id': link.get('data-student_id'),
                        'student_no': link.get('data-student_no'),
                        'name': link.get('data-student_name'),
                        'name_cn': link.get('data-student_cname'),
                        'class_name': link.get('data-class_name'),
                        'class_id': link.get('data-class_id'),
                    }
                    if student['internal_id']:  # 确保有 internal_id
                        students.append(student)
            
            self.log(f"  ✓ 获得 {len(students)} 名学生 (HTTP {resp.status_code})")
            return students
            
        except Exception as e:
            self.log(f"  WARNING: 获取班级 {class_name} 学生列表异常: {e}", "WARN")
            return []
    
    def fetch_all_students(self) -> Dict[str, List[Dict]]:
        """获取所有班级的所有学生
        
        Returns:
            {class_id: [student_list]}
        """
        self.log("=" * 70)
        self.log("开始批量获取所有班级的学生名单")
        self.log("=" * 70)
        
        if not self.class_mapping:
            self.log("ERROR: 班级列表为空，请先登入", "ERROR")
            return {}
        
        self.all_students = {}
        total_students = 0
        success_count = 0
        failed_classes = []
        
        for idx, (class_id, class_name) in enumerate(self.class_mapping.items(), 1):
            progress = f"[{idx}/{len(self.class_mapping)}]"
            self.log(f"{progress} 处理班级: {class_name} (ID: {class_id})")
            
            students = self._get_students_by_class(class_id, class_name)
            if students:
                self.all_students[class_id] = students
                total_students += len(students)
                success_count += 1
                self.log(f"  ✓ 获得 {len(students)} 名学生")
            else:
                self.log(f"  ✗ 无学生数据", "WARN")
                failed_classes.append(class_name)
            
            # 为了避免服务器压力，添加延迟
            time.sleep(0.5)
        
        self.log("=" * 70)
        self.log("批量获取完成")
        self.log(f"  成功班级: {success_count}/{len(self.class_mapping)}")
        self.log(f"  失败班级: {len(failed_classes)}")
        self.log(f"  总学生数: {total_students}")
        self.log("=" * 70)
        
        return self.all_students
    
    def save_to_json(self, output_path: str = None) -> str:
        """将学生数据保存为 JSON 文件
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            保存的文件路径
        """
        if not self.all_students:
            self.log("ERROR: 没有学生数据要保存", "ERROR")
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.data_dir / f"students_{timestamp}.json"
        
        try:
            # 转换为便于查看的格式
            output_data = {
                'download_time': datetime.now().isoformat(),
                'total_classes': len(self.all_students),
                'total_students': sum(len(students) for students in self.all_students.values()),
                'classes': {}
            }
            
            for class_id, students in self.all_students.items():
                class_name = self.class_mapping.get(class_id, f"Class {class_id}")
                output_data['classes'][class_id] = {
                    'name': class_name,
                    'student_count': len(students),
                    'students': students
                }
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            self.log(f"✓ JSON 数据已保存到: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.log(f"ERROR: 保存 JSON 文件失败: {e}", "ERROR")
            return None
    
    def save_to_excel(self, output_path: str = None) -> str:
        """将学生数据保存为 Excel 文件
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            保存的文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.log("ERROR: 需要安装 openpyxl: pip install openpyxl", "ERROR")
            return None
        
        if not self.all_students:
            self.log("ERROR: 没有学生数据要保存", "ERROR")
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = self.data_dir / f"students_{timestamp}.xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "学生名单"
            
            # 设置表头
            headers = ["班级", "学号", "姓名", "中文名", "内部ID"]
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 填充数据
            row = 2
            for class_id, students in self.all_students.items():
                class_name = self.class_mapping.get(class_id, f"Class {class_id}")
                for student in students:
                    ws.append([
                        class_name,
                        student.get('student_no', ''),
                        student.get('name', ''),
                        student.get('name_cn', ''),
                        student.get('internal_id', '')
                    ])
                    row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            wb.save(output_path)
            self.log(f"✓ Excel 数据已保存到: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.log(f"ERROR: 保存 Excel 文件失败: {e}", "ERROR")
            import traceback
            traceback.print_exc()
            return None
    
    def get_summary(self) -> Dict:
        """获取摘要信息
        
        Returns:
            统计信息字典
        """
        total_students = sum(len(students) for students in self.all_students.values())
        return {
            'total_classes': len(self.class_mapping),
            'classes_with_students': len(self.all_students),
            'total_students': total_students,
            'classes': self.class_mapping
        }


def main():
    """主函数 - 使用示例"""
    
    # 配置
    USERNAME = "schhs334"
    PASSWORD = "@Sidan49122"
    DATA_DIR = "C:\\chhsban-acadoc\\download_student"
    
    # 创建获取器（自动选择网络）
    fetcher = SMSStudentListFetcher(data_dir=DATA_DIR)
    
    # 登入
    if not fetcher.login(USERNAME, PASSWORD):
        fetcher.log("ERROR: 登入失败，退出", "ERROR")
        return
    
    # 获取所有班级的学生
    students = fetcher.fetch_all_students()
    
    if not students:
        fetcher.log("ERROR: 未获取到任何学生数据", "ERROR")
        return
    
    # 显示摘要
    summary = fetcher.get_summary()
    print(f"\n{'='*70}")
    print(f"摘要信息:")
    print(f"  班级总数: {summary['total_classes']}")
    print(f"  有学生的班级: {summary['classes_with_students']}")
    print(f"  学生总数: {summary['total_students']}")
    print(f"{'='*70}\n")
    
    # 保存数据
    json_path = fetcher.save_to_json()
    excel_path = fetcher.save_to_excel()
    
    print(f"{'='*70}")
    print(f"前 10 个班级的学生统计:")
    for class_id, class_name in list(summary['classes'].items())[:10]:
        num_students = len(students.get(class_id, []))
        print(f"  {class_name}: {num_students} 名学生")
    if len(summary['classes']) > 10:
        print(f"  ... 等共 {len(summary['classes'])} 个班级")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
