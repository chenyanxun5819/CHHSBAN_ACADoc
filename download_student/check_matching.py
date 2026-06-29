#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import openpyxl

# 从 Excel 中提取所有学号
wb = openpyxl.load_workbook('2026_hostelList.xlsx')
ws = wb.active
excel_students = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    student_id = row[2]  # C 栏
    gender = row[5]      # F 栏
    if student_id:
        excel_students[str(student_id)] = gender

print(f"Excel 学号范围: {min(excel_students.keys())} - {max(excel_students.keys())}")
print(f"Excel 中学号前缀统计:")
prefixes = {}
for sid in excel_students.keys():
    prefix = sid[:2]
    prefixes[prefix] = prefixes.get(prefix, 0) + 1
for prefix in sorted(prefixes.keys()):
    print(f"  {prefix}*****: {prefixes[prefix]} 人")

# 从下载数据中提取所有学号
with open('students_final_20260629_125418.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    download_students = {}
    for student in data['students']:
        sno = student.get('student_no')
        if sno:
            download_students[str(sno)] = student

print(f"\n下载数据学号范围: {min(download_students.keys())} - {max(download_students.keys())}")
print(f"下载数据中学号前缀统计:")
prefixes = {}
for sno in download_students.keys():
    prefix = sno[:2]
    prefixes[prefix] = prefixes.get(prefix, 0) + 1
for prefix in sorted(prefixes.keys()):
    print(f"  {prefix}*****: {prefixes[prefix]} 人")

# 尝试匹配
print(f"\n匹配测试 (前10个下载数据):")
matches = 0
for sno in list(download_students.keys())[:10]:
    if sno in excel_students:
        matches += 1
        print(f"  ✓ {sno} 匹配 - Gender: {excel_students[sno]}")
    else:
        print(f"  ✗ {sno} 未找到 - 在下载数据中: {download_students[sno]['name_en']}")
