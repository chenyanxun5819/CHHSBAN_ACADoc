#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完全从 SMS 网站动态提取数据
1. 获取班级 SELECT 中的所有班级 ID
2. 对每个班级，获取学生数据
3. 去重并统计
"""

import requests
from bs4 import BeautifulSoup
import json
import csv
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl

requests.packages.urllib3.disable_warnings()


class DynamicStudentFetcher:
    def __init__(self):
        self.base_url = None
        self.session = None
        self.all_students = []  # 所有获取的学生记录
        self.unique_students = {}  # 按 student_id 去重后的学生
        self.data_dir = Path("C:/chhsban-acadoc/download_student")
        self.data_dir.mkdir(parents=True, exist_ok=True)
        
        # 读取 Excel 文件的性别/宿舍信息
        self.gender_boarding = self._load_excel_data()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = self.data_dir / f"fetch_dynamic_{timestamp}.log"
    
    def _load_excel_data(self):
        """从 Excel 文件读取性别/宿舍信息，返回学号->性别/宿舍的映射"""
        gender_boarding = {}
        excel_file = self.data_dir / "2026_hostelList.xlsx"
        
        if not excel_file.exists():
            return gender_boarding
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # 读取数据，C 栏是学号，F 栏是性别/宿舍
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                student_id = row[2]  # C 栏
                gender = row[5]      # F 栏
                
                if student_id:
                    gender_boarding[str(student_id)] = gender
            
            return gender_boarding
        except Exception as e:
            print(f"⚠ 读取 Excel 文件出错: {e}")
            return {}
    
    def log(self, msg):
        """输出和记录日志"""
        print(msg)
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(msg + '\n')
    
    def select_network(self):
        """选择可用的网络"""
        networks = {
            'intranet': 'http://192.168.0.6/sms/index.php',
            'internet': 'http://sms.chhsban.edu.my/sms/index.php'
        }
        for name, url in networks.items():
            try:
                resp = requests.get(url, timeout=5, verify=False)
                if resp.status_code == 200:
                    self.base_url = url
                    self.log(f"✓ 连接到 {name}: {url}")
                    return True
            except:
                pass
        self.log("ERROR: 无法连接任何网络")
        return False
    
    def login(self, username, password):
        """登入 SMS 系统"""
        self.session = requests.Session()
        self.session.verify = False
        
        login_url = f"{self.base_url}?r=site/login"
        self.log(f"正在登入...")
        
        resp = self.session.get(login_url)
        login_data = {
            'LoginForm[username]': username,
            'LoginForm[password]': password,
            'login-button': 'login'
        }
        resp = self.session.post(login_url, data=login_data, timeout=15)
        
        if 'login' in resp.url.lower():
            self.log("ERROR: 登入失败")
            return False
        
        self.log("✓ 登入成功")
        return True
    
    def extract_class_list(self):
        """从网页 SELECT 中动态提取班级列表"""
        self.log("\n第1步：从网站提取班级列表...")
        
        activity_page = f"{self.base_url}?r=transaction/studentPerformance/create"
        resp = self.session.get(activity_page, timeout=15)
        
        if resp.status_code != 200:
            self.log(f"ERROR: 获取活动页面失败: HTTP {resp.status_code}")
            return []
        
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # 方法1：查找所有 select，根据名称识别班级列表
        all_selects = soup.find_all('select')
        class_select = None
        
        self.log(f"  页面中共有 {len(all_selects)} 个 SELECT 元素")
        
        for select in all_selects:
            name = select.get('name', '')
            select_id = select.get('id', '')
            options = select.find_all('option')
            
            self.log(f"    - name='{name}', id='{select_id}': {len(options)} 个 option")
            
            # 查找名称为 class_id 的 select
            if name == 'class_id' and len(options) > 10:
                class_select = select
                self.log(f"      ✓ 这是班级 SELECT")
                break
        
        if not class_select:
            self.log("ERROR: 未找到班级 SELECT")
            return []
        
        # 提取班级 ID 和名称
        class_ids = {}  # {class_id: class_name}
        options = class_select.find_all('option')
        
        for option in options:
            class_id = option.get('value', '').strip()
            class_name = option.get_text().strip()
            
            if class_id and class_name:
                class_ids[class_id] = class_name
        
        self.log(f"✓ 从 SELECT 中提取了 {len(class_ids)} 个班级")
        self.log(f"  班级列表: {list(class_ids.keys())[:10]}...")
        
        return class_ids
    
    def fetch_students_for_class(self, class_id, class_name):
        """获取单个班级的学生数据（支持分页）"""
        activity_page = f"{self.base_url}?r=transaction/studentPerformance/create"
        
        students = []
        page = 1
        
        while True:
            try:
                # 使用 id_page 参数支持分页
                params = {'class_id': class_id, 'ajax': 'student-grid', 'id_page': page}
                resp = self.session.get(activity_page, params=params, timeout=15)
                
                if resp.status_code != 200:
                    break
                
                soup = BeautifulSoup(resp.text, 'html.parser')
                page_students = []
                
                # 从 HTML 中提取学生数据
                for link in soup.select('a[data-student_id]'):
                    student = {
                        'student_id': link.get('data-student_id'),
                        'student_no': link.get('data-student_no'),
                        'name_en': link.get('data-student_name'),
                        'name_cn': link.get('data-student_cname'),
                        'real_class_name': link.get('data-class_name'),  # 学生实际班级
                        'real_class_id': link.get('data-class_id'),      # 学生实际班级 ID
                        'input_class_id': class_id,                       # SELECT 中的班级 ID
                        'input_class_name': class_name,                   # SELECT 中的班级名
                    }
                    
                    if student['student_id']:
                        page_students.append(student)
                        self.all_students.append(student)
                
                if not page_students:
                    # 没有数据，结束分页
                    break
                
                students.extend(page_students)
                
                # 检查是否有下一页
                pagination = soup.find('div', class_='pagination')
                if pagination:
                    links = pagination.find_all('a')
                    has_next = False
                    for link in links:
                        href = link.get('href', '')
                        if f'id_page={page + 1}' in href:
                            has_next = True
                            break
                    if not has_next:
                        break
                else:
                    break
                
                page += 1
                
            except Exception as e:
                self.log(f"      分页错误 (页 {page}): {str(e)}")
                break
        
        return students
    
    def fetch_all_students(self, class_ids):
        """获取所有班级的学生"""
        self.log(f"\n第2步：逐一获取每个班级的学生...")
        
        for idx, (class_id, class_name) in enumerate(class_ids.items(), 1):
            print(f"  [{idx:2d}/{len(class_ids)}] {class_name:30s} (ID: {class_id:3s})...", end='', flush=True)
            
            students = self.fetch_students_for_class(class_id, class_name)
            print(f" ✓ {len(students):4d} 条记录")
            
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(f"  [{idx:2d}] {class_name}: {len(students)} 条\n")
        
        self.log(f"\n✓ 总共获取 {len(self.all_students)} 条学生记录")
    
    def deduplicate(self):
        """按 student_id 去重"""
        self.log(f"\n第3步：去重处理...")
        
        self.unique_students = {}
        
        for student in self.all_students:
            sid = student.get('student_id')
            if sid not in self.unique_students:
                self.unique_students[sid] = student
        
        removed = len(self.all_students) - len(self.unique_students)
        self.log(f"  原始记录: {len(self.all_students)}")
        self.log(f"  去重后: {len(self.unique_students)}")
        self.log(f"  删除重复: {removed} ({removed/len(self.all_students)*100:.1f}%)")
        
        return list(self.unique_students.values())
    
    def save_data(self, unique_students):
        """保存数据为 JSON 和 CSV，并添加 Gender/Boarding 信息"""
        self.log(f"\n第4步：保存数据...")
        
        if not unique_students:
            self.log("ERROR: 没有数据")
            return
        
        # 添加 Gender/Boarding 信息
        enriched_students = []
        for student in unique_students:
            student_copy = student.copy()
            student_no = str(student.get('student_no', ''))
            
            # 从 Excel 中查找性别/宿舍信息
            if student_no in self.gender_boarding:
                student_copy['gender_boarding'] = self.gender_boarding[student_no]
            else:
                student_copy['gender_boarding'] = None
            
            enriched_students.append(student_copy)
        
        # 统计班级和学生
        classes_count = len(set(s.get('real_class_name') for s in enriched_students))
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 保存 JSON
        json_file = self.data_dir / f"students_final_{timestamp}.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump({
                'download_time': datetime.now().isoformat(),
                'total_students': len(enriched_students),
                'total_classes': classes_count,
                'students': enriched_students
            }, f, ensure_ascii=False, indent=2)
        self.log(f"✓ 保存 JSON: {json_file} ({json_file.stat().st_size / 1024 / 1024:.2f} MB)")
        
        # 保存 CSV
        csv_file = self.data_dir / f"students_final_{timestamp}.csv"
        with open(csv_file, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=[
                'input_class_name', 'real_class_name', 'student_no', 
                'name_en', 'name_cn', 'student_id', 'gender_boarding'
            ], extrasaction='ignore')
            writer.writeheader()
            writer.writerows(enriched_students)
        self.log(f"✓ 保存 CSV: {csv_file}")
        
        # 班级统计
        self.log(f"\n✓ 班级统计（共 {classes_count} 个班级）:")
        class_stats = defaultdict(int)
        for student in enriched_students:
            class_name = student.get('real_class_name', 'Unknown')
            class_stats[class_name] += 1
        
        for class_name in sorted(class_stats.keys()):
            self.log(f"  {class_name:20s}: {class_stats[class_name]:3d} 名学生")


def main():
    USERNAME = "schhs334"
    PASSWORD = "@Sidan49122"
    
    fetcher = DynamicStudentFetcher()
    
    if not fetcher.select_network():
        return
    
    if not fetcher.login(USERNAME, PASSWORD):
        return
    
    # 第1步：从网站提取班级列表
    class_ids = fetcher.extract_class_list()
    if not class_ids:
        return
    
    # 第2步：获取所有班级的学生
    fetcher.fetch_all_students(class_ids)
    
    # 第3步：去重
    unique_students = fetcher.deduplicate()
    
    # 第4步：保存
    fetcher.save_data(unique_students)
    
    fetcher.log("\n✅ 处理完成！")


if __name__ == "__main__":
    main()
