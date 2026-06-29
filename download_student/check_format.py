#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import openpyxl

# 查看下载的数据格式
with open('students_final_20260629_125418.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
    print('下载数据中的学生:')
    for i, student in enumerate(data['students'][:3], 1):
        print(f"  {i}. student_no={student.get('student_no')}, name_en={student.get('name_en')}, real_class={student.get('real_class_name')}")

# 查看 Excel 数据格式
print('\nExcel 数据中的学生:')
wb = openpyxl.load_workbook('2026_hostelList.xlsx')
ws = wb.active
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True), 1):
    print(f"  {i}. studentID={row[2]}, EngName={row[3]}, ChnName={row[4]}, Gender={row[5]}")

# 统计 Excel 的学生数
excel_count = ws.max_row - 1
print(f'\nExcel 总学生数: {excel_count}')
print(f'下载数据总学生数: {len(data["students"])}')
