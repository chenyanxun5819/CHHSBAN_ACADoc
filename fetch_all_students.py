#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS 系统 - 批量获取所有班级的学生名单
独立脚本，可用于其他项目
"""

import requests
from bs4 import BeautifulSoup
import json
import time
from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime

# 禁用 SSL 警告
requests.packages.urllib3.disable_warnings()


class SMSStudentListFetcher:
    """从 SMS 系统获取所有班级的学生名单"""
    
    def __init__(self, base_url: str = "http://192.168.0.6/sms/index.php"):
        """初始化获取器
        
        Args:
            base_url: SMS 系统基础 URL
        """
        self.base_url = base_url
        self.session = None
        self.class_mapping = {}  # class_id → class_name
        self.all_students = {}   # class_id → [student_list]
        
        # URL 配置
        self.login_url = f"{self.base_url}?r=site/login"
        self.activity_page = f"{self.base_url}?r=transaction/studentPerformance/create"
    
    def _init_session(self):
        """初始化请求会话"""
        if not self.session:
            self.session = requests.Session()
            self.session.verify = False
            self.session.timeout = 30
        return self.session
    
    def login(self, username: str, password: str) -> bool:
        """登入 SMS 系统
        
        Args:
            username: 用户名
            password: 密码
            
        Returns:
            是否登入成功
        """
        try:
            session = self._init_session()
            print(f"[INFO] 开始登入...")
            print(f"       用户: {username}")
            
            # 第1步：GET 登入页面
            print(f"[INFO] 1️⃣  获取登入页面...")
            resp = session.get(self.login_url, timeout=15)
            if resp.status_code != 200:
                print(f"[ERROR] 获取登入页面失败: HTTP {resp.status_code}")
                return False
            print(f"[INFO] ✓ 页面已获取")
            
            # 第2步：POST 登入凭证
            print(f"[INFO] 2️⃣  提交登入凭证...")
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            resp = session.post(self.login_url, data=login_data, timeout=15, allow_redirects=True)
            
            # 检查是否仍在登入页面
            if 'login' in resp.url.lower():
                print(f"[ERROR] 登入失败: 仍停留在登入页面")
                return False
            
            print(f"[INFO] ✓ 登入成功")
            print(f"[INFO] 当前 URL: {resp.url}")
            
            # 第3步：获取班级列表
            print(f"[INFO] 3️⃣  提取班级列表...")
            if self._fetch_class_list():
                print(f"[INFO] ✓ 班级列表已获取 ({len(self.class_mapping)} 个班级)")
            else:
                print(f"[WARNING] 班级列表获取失败，但继续处理")
            
            return True
            
        except Exception as e:
            print(f"[ERROR] 登入异常: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fetch_class_list(self) -> bool:
        """从页面的 SELECT 元素提取班级列表
        
        Returns:
            是否成功提取
        """
        try:
            resp = self.session.get(self.activity_page, timeout=15)
            if resp.status_code != 200:
                print(f"[ERROR] 获取活动页面失败: HTTP {resp.status_code}")
                return False
            
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 查找班级选择器 SELECT 元素
            class_select = soup.select_one('select#StudentPerformanceM_class_id')
            if not class_select:
                print(f"[WARNING] 未找到班级选择器 select#StudentPerformanceM_class_id")
                return False
            
            # 从 SELECT 中提取所有 option
            self.class_mapping = {}
            options = class_select.find_all('option')
            for option in options:
                class_id = option.get('value', '').strip()
                class_name = option.get_text().strip()
                
                # 跳过空白或占位符选项
                if class_id and class_name and class_name != '--- 选择班级 ---':
                    self.class_mapping[class_id] = class_name
            
            print(f"[INFO] ✓ 提取了 {len(self.class_mapping)} 个班级")
            return len(self.class_mapping) > 0
            
        except Exception as e:
            print(f"[ERROR] 提取班级列表异常: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_students_by_class(self, class_id: str, class_name: str) -> List[Dict]:
        """获取指定班级的所有学生
        
        Args:
            class_id: 班级 ID
            class_name: 班级名称
            
        Returns:
            学生列表，每项包含 internal_id, student_no, name, class_id 等
        """
        try:
            # 向活动页面请求班级数据
            params = {'class_id': class_id}
            resp = self.session.get(self.activity_page, params=params, timeout=15)
            
            if resp.status_code != 200:
                print(f"[WARNING] 获取班级 {class_name} (ID: {class_id}) 失败: HTTP {resp.status_code}")
                return []
            
            # 解析 HTML
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 查找所有学生行 - a 标签带 data-student_id 属性
            students = []
            for link in soup.select('a[data-student_id]'):
                student = {
                    'internal_id': link.get('data-student_id'),
                    'student_no': link.get('data-student_no'),
                    'name': link.get('data-student_name'),
                    'name_cn': link.get('data-student_cname'),
                    'class_name': link.get('data-class_name'),
                    'class_id': link.get('data-class_id'),
                }
                if student['internal_id']:  # 确保有 internal_id
                    students.append(student)
            
            return students
            
        except Exception as e:
            print(f"[WARNING] 获取班级 {class_name} 学生列表异常: {e}")
            return []
    
    def fetch_all_students(self) -> Dict[str, List[Dict]]:
        """获取所有班级的所有学生
        
        Returns:
            {class_id: [student_list]}
        """
        print(f"\n{'='*70}")
        print(f"[INFO] 开始批量获取所有班级的学生名单")
        print(f"{'='*70}\n")
        
        if not self.class_mapping:
            print(f"[ERROR] 班级列表为空，请先登入")
            return {}
        
        self.all_students = {}
        total_students = 0
        success_count = 0
        
        for idx, (class_id, class_name) in enumerate(self.class_mapping.items(), 1):
            print(f"[INFO] 处理班级 {idx}/{len(self.class_mapping)}: {class_name} (ID: {class_id})")
            
            students = self._get_students_by_class(class_id, class_name)
            if students:
                self.all_students[class_id] = students
                total_students += len(students)
                success_count += 1
                print(f"       ✓ 获得 {len(students)} 名学生")
            else:
                print(f"       ✗ 无学生数据")
            
            # 为了避免服务器压力，添加延迟
            time.sleep(0.5)
        
        print(f"\n{'='*70}")
        print(f"[INFO] 批量获取完成")
        print(f"       成功班级: {success_count}/{len(self.class_mapping)}")
        print(f"       总学生数: {total_students}")
        print(f"{'='*70}\n")
        
        return self.all_students
    
    def save_to_json(self, output_path: str = None) -> str:
        """将学生数据保存为 JSON 文件
        
        Args:
            output_path: 输出文件路径，默认为当前目录
            
        Returns:
            保存的文件路径
        """
        if not self.all_students:
            print(f"[ERROR] 没有学生数据要保存")
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"sms_students_{timestamp}.json"
        
        try:
            # 转换为便于查看的格式
            output_data = {}
            for class_id, students in self.all_students.items():
                class_name = self.class_mapping.get(class_id, f"Class {class_id}")
                output_data[class_name] = students
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            
            print(f"[INFO] ✓ 数据已保存到: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] 保存 JSON 文件失败: {e}")
            return None
    
    def save_to_excel(self, output_path: str = None) -> str:
        """将学生数据保存为 Excel 文件
        
        Args:
            output_path: 输出文件路径，默认为当前目录
            
        Returns:
            保存的文件路径
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print(f"[ERROR] 需要安装 openpyxl: pip install openpyxl")
            return None
        
        if not self.all_students:
            print(f"[ERROR] 没有学生数据要保存")
            return None
        
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"sms_students_{timestamp}.xlsx"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "学生名单"
            
            # 设置表头
            headers = ["班级", "学号", "姓名", "中文名", "内部ID"]
            ws.append(headers)
            
            # 设置表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 填充数据
            row = 2
            for class_id, students in self.all_students.items():
                class_name = self.class_mapping.get(class_id, f"Class {class_id}")
                for student in students:
                    ws.append([
                        class_name,
                        student.get('student_no', ''),
                        student.get('name', ''),
                        student.get('name_cn', ''),
                        student.get('internal_id', '')
                    ])
                    row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            
            wb.save(output_path)
            print(f"[INFO] ✓ 数据已保存到: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"[ERROR] 保存 Excel 文件失败: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def get_summary(self) -> Dict:
        """获取摘要信息
        
        Returns:
            统计信息字典
        """
        total_students = sum(len(students) for students in self.all_students.values())
        return {
            'total_classes': len(self.class_mapping),
            'classes_with_students': len(self.all_students),
            'total_students': total_students,
            'classes': self.class_mapping
        }


def main():
    """主函数 - 使用示例"""
    
    # 配置
    USERNAME = "your_username"  # 修改为实际用户名
    PASSWORD = "your_password"  # 修改为实际密码
    BASE_URL = "http://192.168.0.6/sms/index.php"
    
    # 创建获取器
    fetcher = SMSStudentListFetcher(base_url=BASE_URL)
    
    # 登入
    if not fetcher.login(USERNAME, PASSWORD):
        print("[ERROR] 登入失败，退出")
        return
    
    # 获取所有班级的学生
    students = fetcher.fetch_all_students()
    
    if not students:
        print("[ERROR] 未获取到任何学生数据")
        return
    
    # 显示摘要
    summary = fetcher.get_summary()
    print(f"\n[INFO] 摘要信息:")
    print(f"       班级总数: {summary['total_classes']}")
    print(f"       有学生的班级: {summary['classes_with_students']}")
    print(f"       学生总数: {summary['total_students']}\n")
    
    # 保存数据
    json_path = fetcher.save_to_json()
    excel_path = fetcher.save_to_excel()
    
    print(f"\n[INFO] 所有班级:")
    for class_id, class_name in summary['classes'].items():
        num_students = len(students.get(class_id, []))
        print(f"       {class_name}: {num_students} 名学生")


if __name__ == "__main__":
    main()
